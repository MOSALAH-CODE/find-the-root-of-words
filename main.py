import sys

import grpc

import zemberek_grpc.language_id_pb2_grpc as z_langid_g
import zemberek_grpc.normalization_pb2_grpc as z_normalization_g
import zemberek_grpc.preprocess_pb2_grpc as z_preprocess_g
import zemberek_grpc.morphology_pb2 as z_morphology
import zemberek_grpc.morphology_pb2_grpc as z_morphology_g

import pandas as pd

from openpyxl import load_workbook
from urllib3.connectionpool import xrange

# The source xlsx file is named as source.xlsx
wb = load_workbook("data_set_1k.xlsx")

ws = wb.active
tweet_column = ws['B']

channel = grpc.insecure_channel('localhost:6789')

langid_stub = z_langid_g.LanguageIdServiceStub(channel)
normalization_stub = z_normalization_g.NormalizationServiceStub(channel)
preprocess_stub = z_preprocess_g.PreprocessingServiceStub(channel)
morphology_stub = z_morphology_g.MorphologyServiceStub(channel)


def analyze(i):
    response = morphology_stub.AnalyzeSentence(z_morphology.SentenceAnalysisRequest(input=i))
    return response


def fix_decode(text):
    """Pass decode."""
    if sys.version_info < (3, 0):
        return text.decode('utf-8')
    else:
        return text


words = []


def run():
    for x in xrange(len(tweet_column)):
        analysis_input = tweet_column[x].value
        if analysis_input is None:
            break
        print(f'{x}. Analysis result for input : ' + fix_decode(analysis_input))
        analysis_result = analyze(analysis_input)
        for a in analysis_result.results:
            best = a.best
            lemmas = best.lemmas[0]
            words.append(lemmas)


if __name__ == '__main__':
    run()
    # print(words)
    unique_words = list(set(words))
    # print(unique_words)
    print("length of all word:", len(words))
    print("length of all word (unique):", len(unique_words))

    df = pd.DataFrame([unique_words], columns=range(len(unique_words)))
    df.to_excel("unique_words.xlsx", index=False, header=False)
